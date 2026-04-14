import re
from datetime import date
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

ARCHIVO = "ESTADISTICAS FUTBEER FC 2026 (22).xlsx"
HOJA_JUGADORES = "Acumulados_Jugadores 2026"
HOJA_BASE = "Base_Partidos 2026"

COLUMNAS_BASE = [
    "Fecha",
    "Partido #",
    "Jugador",
    "Equipo",
    "Goles",
    "Asistencias",
    "Resultado",
    "MVP",
    "Goles Arquero",
    "Captura desde App",  # Nueva columna para indicar si fue ingresado desde la app
]

ARQUEROS_FUTBEER = [
    "Collantes Arquero",
    "Grijalva Arquero",
    "Henry",
    "William Omaña",
    "Diego Ortega",
    "Jesus Arquero Luis A",
    "Luis Angel",
]

LOGO_CANDIDATOS = [
    Path("assets/logo_futbeer.png"),
    Path("assets/logo_futbeer.jpg"),
    Path("assets/logo_futbeer.jpeg"),
    Path("logo_futbeer.png"),
    Path("logo_futbeer.jpg"),
    Path("logo_futbeer.jpeg"),
]

PORTADA_CANDIDATOS = [
    Path("assets/portada_futbeer.png"),
    Path("assets/portada_futbeer.jpg"),
    Path("assets/portada_futbeer.jpeg"),
    Path("portada_futbeer.png"),
    Path("portada_futbeer.jpg"),
    Path("portada_futbeer.jpeg"),
]


def buscar_archivo(candidatos):
    for ruta in candidatos:
        if ruta.exists():
            return ruta
    return None


logo_path = buscar_archivo(LOGO_CANDIDATOS)
portada_path = buscar_archivo(PORTADA_CANDIDATOS)

st.set_page_config(
    page_title="FORMULARIO CAPTURA DE PARTIDO FUTBEER FC",
    page_icon=str(logo_path) if logo_path else "⚽",
    layout="wide",
    initial_sidebar_state="collapsed"
)
<style>
.main {
    background: linear-gradient(180deg, #f7f8fc 0%, #f2effb 100%);
}
.block-container {
    padding-top: 0.8rem;
    padding-bottom: 1.3rem;
    max-width: 1240px;
}
h1, h2, h3 {
    color: #1e2746;
}
.hero-box {
    background: linear-gradient(135deg, #111111 0%, #2a1548 50%, #d8ac3d 100%);
    color: white;
    border-radius: 20px;
    padding: 18px 20px;
    box-shadow: 0 10px 24px rgba(20,20,20,0.16);
    margin-bottom: 12px;
}
.hero-box h1 {
    color: white;
    margin-bottom: 0.25rem;
    font-size: 1.8rem;
}
.hero-box p {
    margin: 0;
    opacity: 0.96;
    font-size: 0.95rem;
}
.section-card {
    background: rgba(255,255,255,0.93);
    border: 1px solid rgba(60,60,100,0.08);
    border-radius: 18px;
    padding: 12px 14px;
    box-shadow: 0 6px 18px rgba(25,30,60,0.05);
    margin-bottom: 10px;
}
.metric-box {
    background: white;
    border-radius: 14px;
    padding: 10px 12px;
    border: 1px solid #ececf5;
    text-align: center;
    box-shadow: 0 4px 12px rgba(20,20,30,0.04);
}
.metric-label {
    font-size: 0.85rem;
    color: #6a7289;
}
.metric-value {
    font-size: 1.35rem;
    font-weight: 700;
    color: #20273e;
}
.team-title-black {
    background: #151515;
    color: white;
    padding: 10px 14px;
    border-radius: 12px;
    margin-bottom: 8px;
    font-weight: 700;
}
.team-title-white {
    background: #ffffff;
    color: #1f2a44;
    padding: 10px 14px;
    border-radius: 12px;
    border: 1px solid #e6e6ef;
    margin-bottom: 8px;
    font-weight: 700;
}
.footer-note {
    color: #5f6780;
    font-size: 0.92rem;
}
.portada-box {
    background: white;
    padding: 8px;
    border-radius: 18px;
    box-shadow: 0 6px 18px rgba(25,30,60,0.08);
    margin-bottom: 12px;
}
.row-player {
    background: #faf9fe;
    border: 1px solid #efecfb;
    border-radius: 10px;
    padding: 6px 8px;
    margin-bottom: 6px;
}
.badge-keeper {
    display: inline-block;
    background: #ffe9a8;
    color: #5e4700;
    border-radius: 999px;
    padding: 2px 8px;
    font-size: 0.7rem;
    font-weight: 700;
    margin-left: 8px;
}
.small-count {
    font-size: 0.85rem;
    color: #5f6780;
}
.stButton > button {
    height: 2.9rem;
    border-radius: 12px;
    font-size: 1rem;
    font-weight: 700;
}
div[data-testid="stTabs"] button {
    font-size: 0.93rem;
}
</style>
""", unsafe_allow_html=True)


def slug(texto: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "_", str(texto).strip().lower())


def titulo_con_logo(texto):
    c1, c2 = st.columns([1, 14])
    with c1:
        if logo_path:
            st.image(str(logo_path), width=40)
    with c2:
        st.subheader(texto)


def cargar_jugadores():
    fuentes = [
        (HOJA_JUGADORES, 1),
        (HOJA_JUGADORES, 0),
        (HOJA_BASE, 0),
    ]

    for hoja, header in fuentes:
        try:
            df = pd.read_excel(ARCHIVO, sheet_name=hoja, header=header)
            df.columns = [str(c).strip() for c in df.columns]

            if "Jugador" in df.columns:
                jugadores = (
                    df["Jugador"]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .tolist()
                )
                jugadores = [j for j in jugadores if j]
                jugadores = sorted(list(dict.fromkeys(jugadores)), key=lambda x: x.lower())
                if jugadores:
                    return jugadores
        except Exception:
            pass

    st.error("No pude cargar la lista de jugadores desde el Excel.")
    st.stop()


def cargar_base():
    try:
        df = pd.read_excel(ARCHIVO, sheet_name=HOJA_BASE)
        df.columns = [str(c).strip() for c in df.columns]
    except Exception:
        df = pd.DataFrame(columns=COLUMNAS_BASE)

    for col in COLUMNAS_BASE:
        if col not in df.columns:
            df[col] = None

    return df[COLUMNAS_BASE].copy()


def siguiente_numero_partido(df_base):
    if df_base.empty:
        return 1

    nums = pd.to_numeric(df_base["Partido #"], errors="coerce").dropna()
    if nums.empty:
        return 1

    return int(nums.max()) + 1


def calcular_ganador(goles_negro: int, goles_blanco: int) -> str:
    if goles_negro > goles_blanco:
        return "Negro"
    if goles_blanco > goles_negro:
        return "Blanco"
    return "Empate"


def resultado_individual(ganador: str, equipo: str) -> str:
    if ganador == "Empate":
        return "E"
    if ganador == equipo:
        return "G"
    return "P"


def append_filas_base(filas):
    wb = load_workbook(ARCHIVO)

    if HOJA_BASE not in wb.sheetnames:
        ws = wb.create_sheet(HOJA_BASE)
        ws.append(COLUMNAS_BASE)
    else:
        ws = wb[HOJA_BASE]
        primera_fila = [c.value for c in ws[1]]

        if ws.max_row == 1 and all(v is None for v in primera_fila):
            ws.delete_rows(1, 1)
            ws.append(COLUMNAS_BASE)
        elif primera_fila != COLUMNAS_BASE:
            raise ValueError(
                "La hoja 'Base_Partidos 2026' tiene encabezados distintos. Revísalos antes de guardar."
            )

    ws = wb[HOJA_BASE]

    for fila in filas:
        fila["Captura desde App"] = "Sí"
        ws.append([
            fila["Fecha"],
            fila["Partido #"],
            fila["Jugador"],
            fila["Equipo"],
            fila["Goles"],
            fila["Asistencias"],
            fila["Resultado"],
            fila["MVP"],
            fila["Goles Arquero"],
            fila["Captura desde App"],  # Guardamos la indicación de que se capturó desde la app
        ])

    wb.save(ARCHIVO)


# Función para seleccionar jugadores y excluir los arqueros del checklist
def seleccionar_jugadores_compacto(nombre_equipo, prefijo, jugadores, arquero_seleccionado):
    # Eliminar arqueros de la lista de jugadores para el checklist
    jugadores_sin_arqueros = [jugador for jugador in jugadores if jugador not in ARQUEROS_FUTBEER]

    if nombre_equipo == "Negro":
        st.markdown('<div class="team-title-black">EQUIPO NEGRO</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="team-title-white">EQUIPO BLANCO</div>', unsafe_allow_html=True)

    if arquero_seleccionado:
        st.success(f"ARQUERO SELECCIONADO: {arquero_seleccionado}")

    buscador = st.text_input(
        f"Buscar jugador en equipo {nombre_equipo}",
        key=f"{prefijo}_buscar"
    ).strip().lower()

    jugadores_filtrados = [
        j for j in jugadores_sin_arqueros if buscador in j.lower()
    ] if buscador else jugadores_sin_arqueros

    cols = st.columns(5)

    for i, jugador in enumerate(jugadores_filtrados):
        key_check = f"{prefijo}_play_{slug(jugador)}"
        with cols[i % 5]:
            etiqueta = jugador
            if jugador == arquero_seleccionado:
                etiqueta = f"{jugador} 🥅"
            st.checkbox(etiqueta, key=key_check)

    participantes = [
        jugador for jugador in jugadores_sin_arqueros
        if st.session_state.get(f"{prefijo}_play_{slug(jugador)}", False)
    ]

    extras_txt = st.text_input(
        f"Agregar jugadores extra {nombre_equipo}, separados por coma",
        key=f"{prefijo}_extras"
    )

    extras = [x.strip() for x in extras_txt.split(",") if x.strip()]

    if arquero_seleccionado and arquero_seleccionado not in participantes and arquero_seleccionado not in extras:
        extras.append(arquero_seleccionado)

    participantes.extend(extras)
    participantes = sorted(list(dict.fromkeys(participantes)), key=lambda x: x.lower())

    st.markdown(
        f'<div class="small-count">Jugadores seleccionados: {len(participantes)}</div>',
        unsafe_allow_html=True
    )

    return participantes

def capturar_estadisticas_equipo_compacto(nombre_equipo, prefijo, participantes, arquero_seleccionado):
    stats = {}

    if not participantes:
        st.info(f"Aún no has marcado jugadores para el equipo {nombre_equipo}.")
        return stats

    with st.expander(f"ESTADÍSTICAS DEL EQUIPO {nombre_equipo.upper()}", expanded=True):
        mitad = (len(participantes) + 1) // 2
        bloques = [participantes[:mitad], participantes[mitad:]]
        col_izq, col_der = st.columns(2)

        for bloque, col in zip(bloques, [col_izq, col_der]):
            with col:
                if bloque:
                    h1, h2, h3 = st.columns([2.3, 1, 1])
                    h1.markdown("**Jugador**")
                    h2.markdown("**Goles**")
                    h3.markdown("**Asist.**")

                for jugador in bloque:
                    badge = ""
                    if jugador == arquero_seleccionado:
                        badge = ' <span class="badge-keeper">ARQUERO</span>'

                    st.markdown(
                        f'<div class="row-player"><strong>{jugador}</strong>{badge}</div>',
                        unsafe_allow_html=True
                    )

                    c1, c2, c3 = st.columns([2.3, 1, 1])
                    c1.caption("")
                    goles = c2.number_input(
                        f"Goles {jugador}",
                        min_value=0,
                        step=1,
                        key=f"{prefijo}_g_{slug(jugador)}",
                        label_visibility="collapsed"
                    )
                    asistencias = c3.number_input(
                        f"Asistencias {jugador}",
                        min_value=0,
                        step=1,
                        key=f"{prefijo}_a_{slug(jugador)}",
                        label_visibility="collapsed"
                    )

                    stats[jugador] = {
                        "goles": int(goles),
                        "asistencias": int(asistencias),
                    }

    return stats


jugadores_excel = cargar_jugadores()
jugadores = sorted(
    list(dict.fromkeys(jugadores_excel + ARQUEROS_FUTBEER)),
    key=lambda x: x.lower()
)

df_base = cargar_base()
numero_partido = siguiente_numero_partido(df_base)

col_logo, col_info = st.columns([1, 6])

with col_logo:
    if logo_path:
        st.image(str(logo_path), use_container_width=True)

with col_info:
    st.markdown("""
    <div class="hero-box">
        <h1>FORMULARIO CAPTURA DE PARTIDO FUTBEER FC</h1>
        <p>Registro rápido del partido, arqueros, MVPs y base lista para Excel.</p>
    </div>
    """, unsafe_allow_html=True)

if portada_path:
    c_port1, c_port2, c_port3 = st.columns([1, 8, 1])
    with c_port2:
        st.markdown('<div class="portada-box">', unsafe_allow_html=True)
        st.image(str(portada_path), use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
else:
    st.warning("No encontré la portada. Guarda la imagen como assets/portada_futbeer.png")

if not logo_path:
    st.warning("No encontré el logo. Guarda la imagen como assets/logo_futbeer.png")

st.markdown('<div class="section-card">', unsafe_allow_html=True)
titulo_con_logo("DATOS GENERALES")

c1, c2 = st.columns(2)
with c1:
    fecha_partido = st.date_input("Fecha", value=date.today())
with c2:
    st.markdown(
        f'<div class="metric-box"><div class="metric-label">Partido #</div><div class="metric-value">{numero_partido}</div></div>',
        unsafe_allow_html=True
    )
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="section-card">', unsafe_allow_html=True)
titulo_con_logo("ARQUEROS")

c_arq1, c_arq2 = st.columns(2)

with c_arq1:
    arquero_negro = st.selectbox("Arquero Negro", [""] + ARQUEROS_FUTBEER, index=0)

with c_arq2:
    arquero_blanco = st.selectbox("Arquero Blanco", [""] + ARQUEROS_FUTBEER, index=0)

if arquero_negro and arquero_blanco and arquero_negro == arquero_blanco:
    st.error("No puedes seleccionar el mismo arquero para ambos equipos.")
    st.stop()

st.markdown('</div>', unsafe_allow_html=True)

tab_negro, tab_blanco = st.tabs(["EQUIPO NEGRO", "EQUIPO BLANCO"])

with tab_negro:
    equipo_negro = seleccionar_jugadores_compacto("Negro", "negro", jugadores, arquero_negro)
    stats_negro = capturar_estadisticas_equipo_compacto("Negro", "negro", equipo_negro, arquero_negro)

with tab_blanco:
    equipo_blanco = seleccionar_jugadores_compacto("Blanco", "blanco", jugadores, arquero_blanco)
    stats_blanco = capturar_estadisticas_equipo_compacto("Blanco", "blanco", equipo_blanco, arquero_blanco)

repetidos = sorted(set(equipo_negro) & set(equipo_blanco), key=lambda x: x.lower())
if repetidos:
    st.error("Hay jugadores repetidos en ambos equipos: " + ", ".join(repetidos))
    st.stop()

goles_negro = sum(stats_negro.get(j, {}).get("goles", 0) for j in equipo_negro)
goles_blanco = sum(stats_blanco.get(j, {}).get("goles", 0) for j in equipo_blanco)
ganador = calcular_ganador(goles_negro, goles_blanco)

st.markdown('<div class="section-card">', unsafe_allow_html=True)
titulo_con_logo("RESUMEN AUTOMÁTICO")

m1, m2, m3 = st.columns(3)

with m1:
    st.markdown(
        f'<div class="metric-box"><div class="metric-label">Goles Negro</div><div class="metric-value">{goles_negro}</div></div>',
        unsafe_allow_html=True
    )
with m2:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
titulo_con_logo("RESUMEN AUTOMÁTICO")

m1, m2, m3 = st.columns(3)

with m1:
    st.markdown(
        f'<div class="metric-box"><div class="metric-label">Goles Negro</div><div class="metric-value">{goles_negro}</div></div>',
        unsafe_allow_html=True
    )
with m2:
    st.markdown(
        f'<div class="metric-box"><div class="metric-label">Goles Blanco</div><div class="metric-value">{goles_blanco}</div></div>',
        unsafe_allow_html=True
    )
with m3:
    st.markdown(
        f'<div class="metric-box"><div class="metric-label">Ganador</div><div class="metric-value">{ganador}</div></div>',
        unsafe_allow_html=True
    )

st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="section-card">', unsafe_allow_html=True)
titulo_con_logo("MVP POR EQUIPO")

c_mvp1, c_mvp2 = st.columns(2)

with c_mvp1:
    mvp_negro = st.selectbox("MVP Equipo Negro", [""] + equipo_negro)

with c_mvp2:
    mvp_blanco = st.selectbox("MVP Equipo Blanco", [""] + equipo_blanco)

st.markdown('</div>', unsafe_allow_html=True)

if st.button("GUARDAR Y GENERAR RESUMEN", use_container_width=True):
    if not arquero_negro:
        st.error("Debes escoger el arquero del equipo Negro.")
        st.stop()

    if not arquero_blanco:
        st.error("Debes escoger el arquero del equipo Blanco.")
        st.stop()

    if not equipo_negro:
        st.error("Falta seleccionar jugadores del equipo Negro.")
        st.stop()

    if not equipo_blanco:
        st.error("Falta seleccionar jugadores del equipo Blanco.")
        st.stop()

    if arquero_negro not in equipo_negro:
        st.error("El arquero del equipo Negro debe quedar incluido entre los jugadores que participaron.")
        st.stop()

    if arquero_blanco not in equipo_blanco:
        st.error("El arquero del equipo Blanco debe quedar incluido entre los jugadores que participaron.")
        st.stop()

    fecha_texto = pd.to_datetime(fecha_partido).strftime("%Y-%m-%d")
    filas_guardar = []
    resumen_negro = []
    resumen_blanco = []

    for jugador in equipo_negro:
        goles = stats_negro[jugador]["goles"]
        asistencias = stats_negro[jugador]["asistencias"]
        es_mvp = "SI" if jugador == mvp_negro else ""
        goles_recibidos = goles_blanco if jugador == arquero_negro else ""

        filas_guardar.append({
            "Fecha": fecha_texto,
            "Partido #": numero_partido,
            "Jugador": jugador,
            "Equipo": "Negro",
            "Goles": goles,
            "Asistencias": asistencias,
            "Resultado": resultado_individual(ganador, "Negro"),
            "MVP": es_mvp,
            "Goles Arquero": goles_recibidos,
        })

        texto = f"* {jugador} | {goles} Goles | {asistencias} Asistencias"
        if jugador == arquero_negro:
            texto += f" | Arquero | {goles_blanco} Goles Recibidos"
        if es_mvp == "SI":
            texto += " | MVP"
        resumen_negro.append(texto)

    for jugador in equipo_blanco:
        goles = stats_blanco[jugador]["goles"]
        asistencias = stats_blanco[jugador]["asistencias"]
        es_mvp = "SI" if jugador == mvp_blanco else ""
        goles_recibidos = goles_negro if jugador == arquero_blanco else ""

        filas_guardar.append({
            "Fecha": fecha_texto,
            "Partido #": numero_partido,
            "Jugador": jugador,
            "Equipo": "Blanco",
            "Goles": goles,
            "Asistencias": asistencias,
            "Resultado": resultado_individual(ganador, "Blanco"),
            "MVP": es_mvp,
            "Goles Arquero": goles_recibidos,
        })

        texto = f"* {jugador} | {goles} Goles | {asistencias} Asistencias"
        if jugador == arquero_blanco:
            texto += f" | Arquero | {goles_negro} Goles Recibidos"
        if es_mvp == "SI":
            texto += " | MVP"
        resumen_blanco.append(texto)

    resumen = f"""FUTBEER FC – CAPTURA DE PARTIDO

Fecha: {fecha_texto}
Partido #: {numero_partido}

Resultado: {goles_negro} - {goles_blanco}
Equipo ganador: {ganador}
MVP Negro: {mvp_negro if mvp_negro else "Sin MVP"}
MVP Blanco: {mvp_blanco if mvp_blanco else "Sin MVP"}

EQUIPO NEGRO
Jugador | Goles | Asistencias
{chr(10).join(resumen_negro)}

EQUIPO BLANCO
Jugador | Goles | Asistencias
{chr(10).join(resumen_blanco)}
"""

    try:
        append_filas_base(filas_guardar)
        st.success("Partido guardado correctamente en Base_Partidos 2026")
        st.text_area("Resumen del partido", resumen, height=420)
    except PermissionError:
        st.error("No pude guardar el Excel. Ciérralo si lo tienes abierto y vuelve a intentar.")
    except Exception as e:
        st.error(f"Ocurrió un error al guardar: {e}")

st.markdown(
    '<p class="footer-note">Pon el logo en assets/logo_futbeer.png y la portada en assets/portada_futbeer.png</p>',
    unsafe_allow_html=True
=======
import re
from datetime import date
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

ARCHIVO = "ESTADISTICAS FUTBEER FC 2026 (22).xlsx"
HOJA_JUGADORES = "Acumulados_Jugadores 2026"
HOJA_BASE = "Base_Partidos 2026"

COLUMNAS_BASE = [
    "Fecha",
    "Partido #",
    "Jugador",
    "Equipo",
    "Goles",
    "Asistencias",
    "Resultado",
    "MVP",
    "Goles Arquero",
    "Captura desde App",  # Nueva columna para indicar si fue ingresado desde la app
]

ARQUEROS_FUTBEER = [
    "Collantes Arquero",
    "Grijalva Arquero",
    "Henry",
    "William Omaña",
    "Diego Ortega",
    "Jesus Arquero Luis A",
    "Luis Angel",
]

LOGO_CANDIDATOS = [
    Path("assets/logo_futbeer.png"),
    Path("assets/logo_futbeer.jpg"),
    Path("assets/logo_futbeer.jpeg"),
    Path("logo_futbeer.png"),
    Path("logo_futbeer.jpg"),
    Path("logo_futbeer.jpeg"),
]

PORTADA_CANDIDATOS = [
    Path("assets/portada_futbeer.png"),
    Path("assets/portada_futbeer.jpg"),
    Path("assets/portada_futbeer.jpeg"),
    Path("portada_futbeer.png"),
    Path("portada_futbeer.jpg"),
    Path("portada_futbeer.jpeg"),
]


def buscar_archivo(candidatos):
    for ruta in candidatos:
        if ruta.exists():
            return ruta
    return None


logo_path = buscar_archivo(LOGO_CANDIDATOS)
portada_path = buscar_archivo(PORTADA_CANDIDATOS)

st.set_page_config(
    page_title="FORMULARIO CAPTURA DE PARTIDO FUTBEER FC",
    page_icon=str(logo_path) if logo_path else "⚽",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
.main {
    background: linear-gradient(180deg, #f7f8fc 0%, #f2effb 100%);
}
.block-container {
    padding-top: 0.8rem;
    padding-bottom: 1.3rem;
    max-width: 1240px;
}
h1, h2, h3 {
    color: #1e2746;
}
.hero-box {
    background: linear-gradient(135deg, #111111 0%, #2a1548 50%, #d8ac3d 100%);
    color: white;
    border-radius: 20px;
    padding: 18px 20px;
    box-shadow: 0 10px 24px rgba(20,20,20,0.16);
    margin-bottom: 12px;
}
.hero-box h1 {
    color: white;
    margin-bottom: 0.25rem;
    font-size: 1.8rem;
}
.hero-box p {
    margin: 0;
    opacity: 0.96;
    font-size: 0.95rem;
}
.section-card {
    background: rgba(255,255,255,0.93);
    border: 1px solid rgba(60,60,100,0.08);
    border-radius: 18px;
    padding: 12px 14px;
    box-shadow: 0 6px 18px rgba(25,30,60,0.05);
    margin-bottom: 10px;
}
.metric-box {
    background: white;
    border-radius: 14px;
    padding: 10px 12px;
    border: 1px solid #ececf5;
    text-align: center;
    box-shadow: 0 4px 12px rgba(20,20,30,0.04);
}
.metric-label {
    font-size: 0.85rem;
    color: #6a7289;
}
.metric-value {
    font-size: 1.35rem;
    font-weight: 700;
    color: #20273e;
}
.team-title-black {
    background: #151515;
    color: white;
    padding: 10px 14px;
    border-radius: 12px;
    margin-bottom: 8px;
    font-weight: 700;
}
.team-title-white {
    background: #ffffff;
    color: #1f2a44;
    padding: 10px 14px;
    border-radius: 12px;
    border: 1px solid #e6e6ef;
    margin-bottom: 8px;
    font-weight: 700;
}
.footer-note {
    color: #5f6780;
    font-size: 0.92rem;
}
.portada-box {
    background: white;
    padding: 8px;
    border-radius: 18px;
    box-shadow: 0 6px 18px rgba(25,30,60,0.08);
    margin-bottom: 12px;
}
.row-player {
    background: #faf9fe;
    border: 1px solid #efecfb;
    border-radius: 10px;
    padding: 6px 8px;
    margin-bottom: 6px;
}
.badge-keeper {
    display: inline-block;
    background: #ffe9a8;
    color: #5e4700;
    border-radius: 999px;
    padding: 2px 8px;
    font-size: 0.7rem;
    font-weight: 700;
    margin-left: 8px;
}
.small-count {
    font-size: 0.85rem;
    color: #5f6780;
}
.stButton > button {
    height: 2.9rem;
    border-radius: 12px;
    font-size: 1rem;
    font-weight: 700;
}
div[data-testid="stTabs"] button {
    font-size: 0.93rem;
}
</style>
""", unsafe_allow_html=True)


def slug(texto: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "_", str(texto).strip().lower())


def titulo_con_logo(texto):
    c1, c2 = st.columns([1, 14])
    with c1:
        if logo_path:
            st.image(str(logo_path), width=40)
    with c2:
        st.subheader(texto)


def cargar_jugadores():
    fuentes = [
        (HOJA_JUGADORES, 1),
        (HOJA_JUGADORES, 0),
        (HOJA_BASE, 0),
    ]

    for hoja, header in fuentes:
        try:
            df = pd.read_excel(ARCHIVO, sheet_name=hoja, header=header)
            df.columns = [str(c).strip() for c in df.columns]

            if "Jugador" in df.columns:
                jugadores = (
                    df["Jugador"]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .tolist()
                )
                jugadores = [j for j in jugadores if j]
                jugadores = sorted(list(dict.fromkeys(jugadores)), key=lambda x: x.lower())
                if jugadores:
                    return jugadores
        except Exception:
            pass

    st.error("No pude cargar la lista de jugadores desde el Excel.")
    st.stop()


def cargar_base():
    try:
        df = pd.read_excel(ARCHIVO, sheet_name=HOJA_BASE)
        df.columns = [str(c).strip() for c in df.columns]
    except Exception:
        df = pd.DataFrame(columns=COLUMNAS_BASE)

    for col in COLUMNAS_BASE:
        if col not in df.columns:
            df[col] = None

    return df[COLUMNAS_BASE].copy()


def siguiente_numero_partido(df_base):
    if df_base.empty:
        return 1

    nums = pd.to_numeric(df_base["Partido #"], errors="coerce").dropna()
    if nums.empty:
        return 1

    return int(nums.max()) + 1


def calcular_ganador(goles_negro: int, goles_blanco: int) -> str:
    if goles_negro > goles_blanco:
        return "Negro"
    if goles_blanco > goles_negro:
        return "Blanco"
    return "Empate"


def resultado_individual(ganador: str, equipo: str) -> str:
    if ganador == "Empate":
        return "E"
    if ganador == equipo:
        return "G"
    return "P"


def append_filas_base(filas):
    wb = load_workbook(ARCHIVO)

    if HOJA_BASE not in wb.sheetnames:
        ws = wb.create_sheet(HOJA_BASE)
        ws.append(COLUMNAS_BASE)
    else:
        ws = wb[HOJA_BASE]
        primera_fila = [c.value for c in ws[1]]

        if ws.max_row == 1 and all(v is None for v in primera_fila):
            ws.delete_rows(1, 1)
            ws.append(COLUMNAS_BASE)
        elif primera_fila != COLUMNAS_BASE:
            raise ValueError(
                "La hoja 'Base_Partidos 2026' tiene encabezados distintos. Revísalos antes de guardar."
            )

    ws = wb[HOJA_BASE]

    for fila in filas:
        fila["Captura desde App"] = "Sí"
        ws.append([
            fila["Fecha"],
            fila["Partido #"],
            fila["Jugador"],
            fila["Equipo"],
            fila["Goles"],
            fila["Asistencias"],
            fila["Resultado"],
            fila["MVP"],
            fila["Goles Arquero"],
            fila["Captura desde App"],  # Guardamos la indicación de que se capturó desde la app
        ])

    wb.save(ARCHIVO)


# Función para seleccionar jugadores y excluir los arqueros del checklist
def seleccionar_jugadores_compacto(nombre_equipo, prefijo, jugadores, arquero_seleccionado):
    # Eliminar arqueros de la lista de jugadores para el checklist
    jugadores_sin_arqueros = [jugador for jugador in jugadores if jugador not in ARQUEROS_FUTBEER]

    if nombre_equipo == "Negro":
        st.markdown('<div class="team-title-black">EQUIPO NEGRO</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="team-title-white">EQUIPO BLANCO</div>', unsafe_allow_html=True)

    if arquero_seleccionado:
        st.success(f"ARQUERO SELECCIONADO: {arquero_seleccionado}")

    buscador = st.text_input(
        f"Buscar jugador en equipo {nombre_equipo}",
        key=f"{prefijo}_buscar"
    ).strip().lower()

    jugadores_filtrados = [
        j for j in jugadores_sin_arqueros if buscador in j.lower()
    ] if buscador else jugadores_sin_arqueros

    cols = st.columns(5)

    for i, jugador in enumerate(jugadores_filtrados):
        key_check = f"{prefijo}_play_{slug(jugador)}"
        with cols[i % 5]:
            etiqueta = jugador
            if jugador == arquero_seleccionado:
                etiqueta = f"{jugador} 🥅"
            st.checkbox(etiqueta, key=key_check)

    participantes = [
        jugador for jugador in jugadores_sin_arqueros
        if st.session_state.get(f"{prefijo}_play_{slug(jugador)}", False)
    ]

    extras_txt = st.text_input(
        f"Agregar jugadores extra {nombre_equipo}, separados por coma",
        key=f"{prefijo}_extras"
    )

    extras = [x.strip() for x in extras_txt.split(",") if x.strip()]

    if arquero_seleccionado and arquero_seleccionado not in participantes and arquero_seleccionado not in extras:
        extras.append(arquero_seleccionado)

    participantes.extend(extras)
    participantes = sorted(list(dict.fromkeys(participantes)), key=lambda x: x.lower())

    st.markdown(
        f'<div class="small-count">Jugadores seleccionados: {len(participantes)}</div>',
        unsafe_allow_html=True
    )

    return participantes

def capturar_estadisticas_equipo_compacto(nombre_equipo, prefijo, participantes, arquero_seleccionado):
    stats = {}

    if not participantes:
        st.info(f"Aún no has marcado jugadores para el equipo {nombre_equipo}.")
        return stats

    with st.expander(f"ESTADÍSTICAS DEL EQUIPO {nombre_equipo.upper()}", expanded=True):
        mitad = (len(participantes) + 1) // 2
        bloques = [participantes[:mitad], participantes[mitad:]]
        col_izq, col_der = st.columns(2)

        for bloque, col in zip(bloques, [col_izq, col_der]):
            with col:
                if bloque:
                    h1, h2, h3 = st.columns([2.3, 1, 1])
                    h1.markdown("**Jugador**")
                    h2.markdown("**Goles**")
                    h3.markdown("**Asist.**")

                for jugador in bloque:
                    badge = ""
                    if jugador == arquero_seleccionado:
                        badge = ' <span class="badge-keeper">ARQUERO</span>'

                    st.markdown(
                        f'<div class="row-player"><strong>{jugador}</strong>{badge}</div>',
                        unsafe_allow_html=True
                    )

                    c1, c2, c3 = st.columns([2.3, 1, 1])
                    c1.caption("")
                    goles = c2.number_input(
                        f"Goles {jugador}",
                        min_value=0,
                        step=1,
                        key=f"{prefijo}_g_{slug(jugador)}",
                        label_visibility="collapsed"
                    )
                    asistencias = c3.number_input(
                        f"Asistencias {jugador}",
                        min_value=0,
                        step=1,
                        key=f"{prefijo}_a_{slug(jugador)}",
                        label_visibility="collapsed"
                    )

                    stats[jugador] = {
                        "goles": int(goles),
                        "asistencias": int(asistencias),
                    }

    return stats


jugadores_excel = cargar_jugadores()
jugadores = sorted(
    list(dict.fromkeys(jugadores_excel + ARQUEROS_FUTBEER)),
    key=lambda x: x.lower()
)

df_base = cargar_base()
numero_partido = siguiente_numero_partido(df_base)

col_logo, col_info = st.columns([1, 6])

with col_logo:
    if logo_path:
        st.image(str(logo_path), use_container_width=True)

with col_info:
    st.markdown("""
    <div class="hero-box">
        <h1>FORMULARIO CAPTURA DE PARTIDO FUTBEER FC</h1>
        <p>Registro rápido del partido, arqueros, MVPs y base lista para Excel.</p>
    </div>
    """, unsafe_allow_html=True)

if portada_path:
    c_port1, c_port2, c_port3 = st.columns([1, 8, 1])
    with c_port2:
        st.markdown('<div class="portada-box">', unsafe_allow_html=True)
        st.image(str(portada_path), use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
else:
    st.warning("No encontré la portada. Guarda la imagen como assets/portada_futbeer.png")

if not logo_path:
    st.warning("No encontré el logo. Guarda la imagen como assets/logo_futbeer.png")

st.markdown('<div class="section-card">', unsafe_allow_html=True)
titulo_con_logo("DATOS GENERALES")

c1, c2 = st.columns(2)
with c1:
    fecha_partido = st.date_input("Fecha", value=date.today())
with c2:
    st.markdown(
        f'<div class="metric-box"><div class="metric-label">Partido #</div><div class="metric-value">{numero_partido}</div></div>',
        unsafe_allow_html=True
    )
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="section-card">', unsafe_allow_html=True)
titulo_con_logo("ARQUEROS")

c_arq1, c_arq2 = st.columns(2)

with c_arq1:
    arquero_negro = st.selectbox("Arquero Negro", [""] + ARQUEROS_FUTBEER, index=0)

with c_arq2:
    arquero_blanco = st.selectbox("Arquero Blanco", [""] + ARQUEROS_FUTBEER, index=0)

if arquero_negro and arquero_blanco and arquero_negro == arquero_blanco:
    st.error("No puedes seleccionar el mismo arquero para ambos equipos.")
    st.stop()

st.markdown('</div>', unsafe_allow_html=True)

tab_negro, tab_blanco = st.tabs(["EQUIPO NEGRO", "EQUIPO BLANCO"])

with tab_negro:
    equipo_negro = seleccionar_jugadores_compacto("Negro", "negro", jugadores, arquero_negro)
    stats_negro = capturar_estadisticas_equipo_compacto("Negro", "negro", equipo_negro, arquero_negro)

with tab_blanco:
    equipo_blanco = seleccionar_jugadores_compacto("Blanco", "blanco", jugadores, arquero_blanco)
    stats_blanco = capturar_estadisticas_equipo_compacto("Blanco", "blanco", equipo_blanco, arquero_blanco)

repetidos = sorted(set(equipo_negro) & set(equipo_blanco), key=lambda x: x.lower())
if repetidos:
    st.error("Hay jugadores repetidos en ambos equipos: " + ", ".join(repetidos))
    st.stop()

goles_negro = sum(stats_negro.get(j, {}).get("goles", 0) for j in equipo_negro)
goles_blanco = sum(stats_blanco.get(j, {}).get("goles", 0) for j in equipo_blanco)
ganador = calcular_ganador(goles_negro, goles_blanco)

st.markdown('<div class="section-card">', unsafe_allow_html=True)
titulo_con_logo("RESUMEN AUTOMÁTICO")

m1, m2, m3 = st.columns(3)

with m1:
    st.markdown(
        f'<div class="metric-box"><div class="metric-label">Goles Negro</div><div class="metric-value">{goles_negro}</div></div>',
        unsafe_allow_html=True
    )
with m2:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
titulo_con_logo("RESUMEN AUTOMÁTICO")

m1, m2, m3 = st.columns(3)

with m1:
    st.markdown(
        f'<div class="metric-box"><div class="metric-label">Goles Negro</div><div class="metric-value">{goles_negro}</div></div>',
        unsafe_allow_html=True
    )
with m2:
    st.markdown(
        f'<div class="metric-box"><div class="metric-label">Goles Blanco</div><div class="metric-value">{goles_blanco}</div></div>',
        unsafe_allow_html=True
    )
with m3:
    st.markdown(
        f'<div class="metric-box"><div class="metric-label">Ganador</div><div class="metric-value">{ganador}</div></div>',
        unsafe_allow_html=True
    )

st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="section-card">', unsafe_allow_html=True)
titulo_con_logo("MVP POR EQUIPO")

c_mvp1, c_mvp2 = st.columns(2)

with c_mvp1:
    mvp_negro = st.selectbox("MVP Equipo Negro", [""] + equipo_negro)

with c_mvp2:
    mvp_blanco = st.selectbox("MVP Equipo Blanco", [""] + equipo_blanco)

st.markdown('</div>', unsafe_allow_html=True)

if st.button("GUARDAR Y GENERAR RESUMEN", use_container_width=True):
    if not arquero_negro:
        st.error("Debes escoger el arquero del equipo Negro.")
        st.stop()

    if not arquero_blanco:
        st.error("Debes escoger el arquero del equipo Blanco.")
        st.stop()

    if not equipo_negro:
        st.error("Falta seleccionar jugadores del equipo Negro.")
        st.stop()

    if not equipo_blanco:
        st.error("Falta seleccionar jugadores del equipo Blanco.")
        st.stop()

    if arquero_negro not in equipo_negro:
        st.error("El arquero del equipo Negro debe quedar incluido entre los jugadores que participaron.")
        st.stop()

    if arquero_blanco not in equipo_blanco:
        st.error("El arquero del equipo Blanco debe quedar incluido entre los jugadores que participaron.")
        st.stop()

    fecha_texto = pd.to_datetime(fecha_partido).strftime("%Y-%m-%d")
    filas_guardar = []
    resumen_negro = []
    resumen_blanco = []

    for jugador in equipo_negro:
        goles = stats_negro[jugador]["goles"]
        asistencias = stats_negro[jugador]["asistencias"]
        es_mvp = "SI" if jugador == mvp_negro else ""
        goles_recibidos = goles_blanco if jugador == arquero_negro else ""

        filas_guardar.append({
            "Fecha": fecha_texto,
            "Partido #": numero_partido,
            "Jugador": jugador,
            "Equipo": "Negro",
            "Goles": goles,
            "Asistencias": asistencias,
            "Resultado": resultado_individual(ganador, "Negro"),
            "MVP": es_mvp,
            "Goles Arquero": goles_recibidos,
        })

        texto = f"* {jugador} | {goles} Goles | {asistencias} Asistencias"
        if jugador == arquero_negro:
            texto += f" | Arquero | {goles_blanco} Goles Recibidos"
        if es_mvp == "SI":
            texto += " | MVP"
        resumen_negro.append(texto)

    for jugador in equipo_blanco:
        goles = stats_blanco[jugador]["goles"]
        asistencias = stats_blanco[jugador]["asistencias"]
        es_mvp = "SI" if jugador == mvp_blanco else ""
        goles_recibidos = goles_negro if jugador == arquero_blanco else ""

        filas_guardar.append({
            "Fecha": fecha_texto,
            "Partido #": numero_partido,
            "Jugador": jugador,
            "Equipo": "Blanco",
            "Goles": goles,
            "Asistencias": asistencias,
            "Resultado": resultado_individual(ganador, "Blanco"),
            "MVP": es_mvp,
            "Goles Arquero": goles_recibidos,
        })

        texto = f"* {jugador} | {goles} Goles | {asistencias} Asistencias"
        if jugador == arquero_blanco:
            texto += f" | Arquero | {goles_negro} Goles Recibidos"
        if es_mvp == "SI":
            texto += " | MVP"
        resumen_blanco.append(texto)

    resumen = f"""FUTBEER FC – CAPTURA DE PARTIDO

Fecha: {fecha_texto}
Partido #: {numero_partido}

Resultado: {goles_negro} - {goles_blanco}
Equipo ganador: {ganador}
MVP Negro: {mvp_negro if mvp_negro else "Sin MVP"}
MVP Blanco: {mvp_blanco if mvp_blanco else "Sin MVP"}

EQUIPO NEGRO
Jugador | Goles | Asistencias
{chr(10).join(resumen_negro)}

EQUIPO BLANCO
Jugador | Goles | Asistencias
{chr(10).join(resumen_blanco)}
"""

    try:
        append_filas_base(filas_guardar)
        st.success("Partido guardado correctamente en Base_Partidos 2026")
        st.text_area("Resumen del partido", resumen, height=420)
    except PermissionError:
        st.error("No pude guardar el Excel. Ciérralo si lo tienes abierto y vuelve a intentar.")
    except Exception as e:
        st.error(f"Ocurrió un error al guardar: {e}")

st.markdown(
    '<p class="footer-note">Pon el logo en assets/logo_futbeer.png y la portada en assets/portada_futbeer.png</p>',
    unsafe_allow_html=True
>>>>>>> b757b6fec8d6c017e25e03293f5c260a630d9772
)